from tkinter import *
from tkinter import ttk
import pandas as pd
from openpyxl import load_workbook
from datetime import date
from functools import partial

today = date.today()

file1=pd.ExcelFile("C:\\Users\\USER\\Desktop\\seating.xlsx")

window=Tk()
v=StringVar()
name=StringVar()
regno=StringVar()
emailid=StringVar()
dept=StringVar()
username=StringVar()
password=StringVar()
window.geometry('400x400')
window.configure(background = 'lightblue2')
button=[]
entry=[]
marks=[]
def clear():
    lis=window.pack_slaves()
    for l in lis:
        l.destroy()
    lis2=window.place_slaves()
    for l in lis2:
        l.destroy()
    v.set("CHOOSE ONE")
    name.set("")
    regno.set("")
    emailid.set("")
    dept.set("")

def set_this(i,j):
    wb = load_workbook("C:\\Users\\USER\\Desktop\\seating.xlsx")
    sheets = wb.sheetnames
    s1 = wb[sheets[j]]
    if s1.cell(1,file1.parse(j).shape[1]+1).value==None:
        s1.cell(1, file1.parse(j).shape[1] + 1).value=today.strftime("%b-%d-%Y")
    s1.cell(i+2, file1.parse(j).shape[1]+1).value = 1
    wb.save("C:\\Users\\USER\\Desktop\\seating.xlsx")

def viewclass(nm,p):
    clear()
    j=0
    for snname in file1.sheet_names:
        if snname==nm:
            c=file1.parse(j)
            break
        j+=1

    i=0
    Label(window, text="", bg="lightblue2").pack(pady=20)
    canvas = Canvas(window,bg="lightblue2",width=500,height=800,highlightbackground="lightblue2",relief="flat")
    scrollbar=Scrollbar(canvas, orient=VERTICAL, command=canvas.yview)
    frame = Frame(canvas,bg="lightblue2",relief="flat")
    frame.grid(row=0,column=0)
    ws = window.winfo_screenwidth()
    hs = window.winfo_screenheight()
    w = 1000
    h = 1000
    x = (ws / 2) - (w / 2)
    y = (hs / 2) - (h / 2)
    for index, row in c.iterrows():
        label=Label(frame, text=row['Name'],font=16, bg="lightblue2").grid(column = 20, row= 30 + i*30)
        i = i + 1
    i=0
    for i in range(c.shape[0]):
         button.append(Button(frame, font=20,text='Present',width=20, bg="snow",command=lambda i=i: set_this(i,j)))
         button[i].grid(column = 200, row = 30 + i*30)
    Button(window, text="BACK", command=lambda: updateatt(p), bg="deepskyblue3").place(x=0, y=0)
    canvas.create_window(0, 0, anchor='nw', window=frame)
    canvas.update_idletasks()
    canvas.configure(scrollregion=canvas.bbox('all'),yscrollcommand=scrollbar.set)
    canvas.pack(fill='both', expand=True, side='left')
    scrollbar.pack(fill='y', side='right')

def updateatt(i):
    clear()
    course=[]
    Button(window, text="BACK", command=welcomeT, font=4, height=1, bg="deepskyblue3").place(x=0, y=0)
    Label(window, text="", bg="lightblue2").pack(pady=20)
    wb = load_workbook("C:\\Users\\USER\\Desktop\\teacherdatabase.xlsx")
    sheets = wb.sheetnames
    s1 = wb[sheets[0]]
    classes=[]
    j=s1.max_column
    for k in range(4,j+1):
        classes.append(str(s1.cell(i,k).value))

    for sname in file1.sheet_names:
        if sname in classes:
            course.append(sname)

    comboExample = ttk.Combobox(window,values=course,state='readonly',width=20,textvariable=v).pack()
    Button(window, text="FIND", command=lambda: viewclass(v.get(),i)).pack(padx=10,pady=10)



def entermarks(nm,p):
    clear()
    Label(window, text='Submit',bg="deepskyblue3",font=("britannic bold",25),relief="solid",highlightthickness=16,highlightbackground='#222').pack(pady=10)
    j = 0
    for snname in file1.sheet_names:
        if snname == nm:
            c = file1.parse(j)
            break
        j += 1

    i = 0
    Label(window, text="", bg="lightblue2").pack(pady=20)
    canvas = Canvas(window, bg="lightblue2",width=500,height=800,highlightbackground="lightblue2", relief="flat")
    scrollbar = Scrollbar(canvas, orient=VERTICAL, command=canvas.yview)
    frame = Frame(canvas, bg="lightblue2", relief="flat")
    frame.grid(row=0, column=0)
    ws = window.winfo_screenwidth()
    hs = window.winfo_screenheight()
    w = 1000
    h = 1000
    x = (ws / 2) - (w / 2)
    y = (hs / 2) - (h / 2)


    '''for i in range(c.shape[0]):
       Label(frame,text=c.iloc[i,0],font=16, bg="lightblue2").grid(row=i,column=1)
       entry.append(Entry(frame, bg='snow', width=20))
       entry[i].grid(row=i, column=10)
       entry[i].bind('<Return>', partial(action,i,j))
    '''

    entries=[]
    for i in range(c.shape[0]):
        Label(frame,text=c.iloc[i,0],font=16, bg="lightblue2").grid(row=i,column=1)
        en = Entry(frame, bg='snow', width=20)
        en.grid(row=i, column=10)
        entries.append(en)

    
            
    def hallo():
        i=0
        for entry in entries:
            text=entry.get()
            wb = load_workbook("C:\\Users\\USER\\Desktop\\seating.xlsx")
            sheets = wb.sheetnames
            s1 = wb[sheets[j]]
            s1.cell(i+2,5).value = text
            i+=1
            wb.save("C:\\Users\\USER\\Desktop\\seating.xlsx")

    button=Button(window,text="Store",command=hallo).pack()
    
    Button(window, text="BACK", command=lambda: updatemarks(p), bg="deepskyblue3").place(x=0, y=0)
    
    canvas.create_window(0, 0, anchor='nw', window=frame)
    canvas.update_idletasks()
    canvas.configure(scrollregion=canvas.bbox('all'), yscrollcommand=scrollbar.set)
    canvas.pack(fill='both', expand=True, side='left')
    scrollbar.pack(fill='y', side='right')

def updatemarks(i):
    clear()
    course=[]
    wb = load_workbook("C:\\Users\\USER\\Desktop\\teacherdatabase.xlsx")
    Button(window, text="BACK", command=welcomeT, font=4, height=1, bg="deepskyblue3").place(x=0, y=0)
    Label(window, text="", bg="lightblue2").pack(pady=20)
    sheets = wb.sheetnames
    s1 = wb[sheets[0]]
    classes=[]
    j=s1.max_column
    for k in range(4,j+1):
        classes.append(str(s1.cell(i,k).value))

    for sname in file1.sheet_names:
        if sname in classes:
            course.append(sname)
    comboExample = ttk.Combobox(window,values=course,state='readonly',width=20,textvariable=v).pack()
    Button(window, text="FIND", command=lambda: entermarks(v.get(),i)).pack(padx=10,pady=10)

def viewstudents(nm,p):
    clear()
    j = 0
    for snname in file1.sheet_names:
        if snname == nm:
            c = file1.parse(j)
            break
        j += 1

    i = 0
    Label(window, text="", bg="lightblue2").pack(pady=20)
    canvas = Canvas(window, bg="lightblue2", width=500, height=800, highlightbackground="lightblue2", relief="flat")
    scrollbar = Scrollbar(canvas, orient=VERTICAL, command=canvas.yview)
    frame = Frame(canvas, bg="lightblue2", relief="flat")
    frame.grid(row=0, column=0)
    ws = window.winfo_screenwidth()
    hs = window.winfo_screenheight()
    w = 1000
    h = 1000
    x = (ws / 2) - (w / 2)
    y = (hs / 2) - (h / 2)
    i = 0
    Label(frame, text="NAME", font=("britannic bold",20), width=13, bg="snow").grid(row=0, column=1)
    Label(frame, text="EMAIL ID", font=("britannic bold",20), width=26, bg="snow").grid(row=0, column=3)
    Label(frame, text="REGISTER NO", font=("britannic bold",20), width=13, bg="snow").grid(row=0, column=5)
    for i in range(c.shape[0]):
        Label(frame,text=c.iloc[i,0],font=16,width=20,height=1, bg="snow").grid(row=i+1,column=1)
        Label(frame, text=c.iloc[i, 1], font=16,width=40,height=1, bg="snow").grid(row=i+1, column=3)
        Label(frame, text=c.iloc[i, 2], font=16,width=20,height=1,bg="snow").grid(row=i+1, column=5)
    Button(window, text="BACK", command=lambda: viewstu(p), bg="deepskyblue3").place(x=0, y=0)
    canvas.create_window(0, 0, anchor='nw', window=frame)
    canvas.update_idletasks()
    canvas.configure(scrollregion=canvas.bbox('all'), yscrollcommand=scrollbar.set)
    canvas.pack(fill='both', expand=True, side='left')
    scrollbar.pack(fill='y', side='right')

def viewstu(i):
    clear()
    course=[]
    wb = load_workbook("C:\\Users\\USER\\Desktop\\teacherdatabase.xlsx")
    Button(window, text="BACK", command=welcomeT, font=4, height=1, bg="deepskyblue3").place(x=0, y=0)
    Label(window, text="", bg="lightblue2").pack(pady=20)
    sheets = wb.sheetnames
    s1 = wb[sheets[0]]
    classes=[]
    j=s1.max_column
    for k in range(4,j+1):
        classes.append(str(s1.cell(i,k).value))

    for sname in file1.sheet_names:
        if sname in classes:
            course.append(sname)
    comboExample = ttk.Combobox(window,values=course,state='readonly',width=20,textvariable=v).pack()
    Button(window, text="FIND", command=lambda: viewstudents(v.get(),i)).pack(padx=10,pady=10)


def welcomeT():
    un = username.get()
    passw = password.get()
    wb = load_workbook("C:\\Users\\USER\\Desktop\\teacherdatabase.xlsx")
    sheets = wb.sheetnames
    s1 = wb[sheets[0]]
    j = 0

    for i in range(1, s1.max_row + 1):
        if (un == s1.cell(i, 2).value):
            j = j + 1
            if (passw == s1.cell(i, 3).value):
                clear()
                Button(window, text="BACK", command=LoginT, font=6, bg="deepskyblue3",relief="groove").place(x=0,y=0)
                Label(window, text="", bg="lightblue2").pack(pady=10)
                Label(window,text='Welcome  teacher', bg="deepskyblue3",font=("britannic bold",25),relief="solid",highlightthickness=16,highlightbackground='#222').pack(pady=10)
                Label(window, text="", bg="lightblue2").pack(pady=5)
                Button(window, text="UPDATE ATTENDANCE", command=lambda : updateatt(i), font=6, bg="deepskyblue3").pack(padx=10, pady=10)
                Button(window, text="ENTER MARKS", command=lambda:updatemarks(i),font=16, bg="deepskyblue3").pack(padx=10, pady=10)
                Button(window, text="VIEW", command=lambda:viewstu(i),font=16, bg="deepskyblue3").pack(padx=10, pady=10)
                break

            else:
                Label(window, text="*INCORRECT PASSWORD", bg="snow", foreground="red").pack(pady=20)
                break
    if j == 0:
        Label(window, text="*INCORRECT CREDENTIALS", bg="snow", foreground="red").pack(pady=20)

def StudentAtt(un):
    j = 0
    clear()
    Button(window, text="BACK", command=welcomeS, font=6, bg="deepskyblue3", relief="groove").place(x=0, y=0)
    Label(window, text="", bg="lightblue2").pack(pady=10)
    wb = load_workbook("C:\\Users\\USER\\Desktop\\seating.xlsx")
    sheets = wb.sheetnames
    for sname in file1.sheet_names:
        c = file1.parse(j)
        #print(c)
        for index, row in c.iterrows():
            if(row['Register no.']==un):
                s1=wb[sheets[j]]
                x=c.shape[1]
                sum=0
                for i in range(5,x):
                    sum=sum+int(s1.cell(index+2,i+1).value)
                Label(window, text="Attendance in class "+ sname+" is "+str(sum)+" / "+str(x-5), font=("britannic bold",20),bg="white").pack(pady=20)


        j += 1
def Studentmarks(un):
    j = 0
    clear()
    Button(window, text="BACK", command=welcomeS, font=6, bg="deepskyblue3", relief="groove").place(x=0, y=0)
    Label(window, text="", bg="lightblue2").pack(pady=10)
    wb = load_workbook("C:\\Users\\USER\\Desktop\\seating.xlsx")
    sheets = wb.sheetnames
    for sname in file1.sheet_names:
        c = file1.parse(j)
        # print(c)
        for index, row in c.iterrows():
            if (row['Register no.'] == un):
                s1 = wb[sheets[j]]
                x = c.shape[1]
                sum = 0
                #for i in range(4, x):
                    #sum = sum + int(s1.cell(index + 2, i + 1).value)
                Label(window, text="Marks in "+str(s1.cell(index + 2, 4).value) + sname + " is " + str(s1.cell(index + 2, 5).value),
                      font=("britannic bold", 20), bg="white").pack(pady=20)

        j += 1

def welcomeS():
    un=username.get()
    passw=password.get()
    wb = load_workbook("C:\\Users\\USER\\Desktop\\teacherdatabase.xlsx")
    sheets = wb.sheetnames
    s1 = wb[sheets[1]]
    j=0

    for i in range (1,s1.max_row+1):
        if(un==s1.cell(i,2).value):
            j=j+1
            if(passw==s1.cell(i,3).value):
                clear()
                Button(window, text="BACK", command=LoginS, font=6, bg="deepskyblue3", relief="groove").place(x=0, y=0)
                Label(window, text="", bg="lightblue2").pack(pady=10)
                Label(window, text='Name: '+s1.cell(i,1).value, bg="deepskyblue3", font=("britannic bold", 20), relief="solid", highlightthickness=16, highlightbackground='#222').pack(pady=10)
                Label(window, text='Register No. ' + s1.cell(i, 2).value, bg="deepskyblue3", font=("britannic bold", 20),relief="solid",highlightthickness=16, highlightbackground='#222').pack(pady=10)
                #Label(window, text='Department: ' + s1.cell(i, 4).value, bg="deepskyblue3", font=("britannic bold", 20),relief="solid",highlightthickness=16, highlightbackground='#222').pack(pady=10)
                Label(window, text="", bg="lightblue2").pack(pady=5)


                Button(window, text="VIEW ATTENDANCE", command=lambda: StudentAtt(un), font=16, bg="deepskyblue3").pack(padx=10,
                                                                                                           pady=10)
                Button(window, text="VIEW MARKS",command=lambda: Studentmarks(un), font=16, bg="deepskyblue3").pack(padx=10, pady=10)
                break

            else:
                Label(window, text="*INCORRECT PASSWORD", bg="snow", foreground="red").pack(pady=20)
                break
    if j==0:
        Label(window, text="*INCORRECT CREDENTIALS", bg="snow", foreground="red").pack(pady=20)

def viewAllStudents():
    clear()
    Button(window, text="BACK", command=viewA, font=4,height=1, bg="deepskyblue3",relief="groove").place(x=0, y=0)
    wb = load_workbook("C:\\Users\\USER\\Desktop\\teacherdatabase.xlsx")
    sheets = wb.sheetnames
    s1 = wb[sheets[1]]
    i = 0
    Label(window, text="", bg="lightblue2").pack(pady=20)
    canvas = Canvas(window, bg="lightblue2", width=500, height=800, highlightbackground="lightblue2", relief="flat")
    scrollbar = Scrollbar(canvas, orient=VERTICAL, command=canvas.yview)
    frame = Frame(canvas, bg="lightblue2", relief="flat")
    frame.grid(row=0, column=0)
    ws = window.winfo_screenwidth()
    hs = window.winfo_screenheight()
    w = 1000
    h = 1000
    x = (ws / 2) - (w / 2)
    y = (hs / 2) - (h / 2)
    i = 0
    Label(frame, text="NAME", font=("britannic bold", 20) ,height=1, width=15, bg="snow").grid(row=0, column=1)
    Label(frame, text="REGISTER NO", font=("britannic bold", 20) ,height=1, width=15, bg="snow").grid(row=0, column=3)
    for i in range(2, s1.max_row + 1):
        Label(frame, text=s1.cell(i, 1).value, height=1, width=20, font=("book antiqua", 12)).grid(row=i, column=1)
        Label(frame, text=s1.cell(i, 2).value, height=1, width=20, font=("book antiqua", 12)).grid(row=i, column=3)

        i = i + 1
    canvas.create_window(0, 0, anchor='nw', window=frame)
    canvas.update_idletasks()
    canvas.configure(scrollregion=canvas.bbox('all'), yscrollcommand=scrollbar.set)
    canvas.pack(fill='both', expand=True, side='left')
    scrollbar.pack(fill='y', side='right')
def attbyadmin(nm):
    clear()
    j = 0
    wb = load_workbook("C:\\Users\\USER\\Desktop\\seating.xlsx")
    sheets = wb.sheetnames
    for snname in file1.sheet_names:
        if snname == nm:
            c = file1.parse(j)
            s1 = wb[sheets[j]]
            ZZ = c.shape[1]
            break
        j += 1

    i = 0
    Label(window, text="", bg="lightblue2").pack(pady=20)
    canvas = Canvas(window, bg="lightblue2", width=500, height=800, highlightbackground="lightblue2", relief="flat")
    scrollbar = Scrollbar(canvas, orient=VERTICAL, command=canvas.yview)
    frame = Frame(canvas, bg="lightblue2", relief="flat")
    frame.grid(row=0, column=0)
    ws = window.winfo_screenwidth()
    hs = window.winfo_screenheight()
    w = 1000
    h = 1000
    x = (ws / 2) - (w / 2)
    y = (hs / 2) - (h / 2)
    Label(frame, text="NAME", font=("britannic bold", 20) ,height=1, width=15, bg="snow").grid(row=0, column=1)
    Label(frame, text="ATTENDENCE", font=("britannic bold", 20) ,height=1, width=15, bg="snow").grid(row=0, column=3)
    i = 1
    for index, row in c.iterrows():
        #label = Label(frame, text=row['Name'], font=16, bg="lightblue2").place(x=0, y=30 + i * 30)
        sum=0
        for M in range(5, ZZ):
            sum = sum + int(s1.cell(index + 2, M + 1).value)
        Label(frame, text=str(s1.cell(index+2,1).value).upper() +" : " ,font= 20, bg="lightblue2").grid(row=i,column=1)
        Label(frame, text=str(sum) + " / " + str(ZZ - 5), font=20,bg="lightblue2").grid(row=i, column=3)

        i = i + 1

    Button(window, text="BACK", command=viewA, bg="deepskyblue3").place(x=0, y=0)
    canvas.create_window(0, 0, anchor='nw', window=frame)
    canvas.update_idletasks()
    canvas.configure(scrollregion=canvas.bbox('all'), yscrollcommand=scrollbar.set)
    canvas.pack(fill='both', expand=True, side='left')
    scrollbar.pack(fill='y', side='right')

def viewmarksbyadmin(nm):
    clear()
    j = 0
    wb = load_workbook("C:\\Users\\USER\\Desktop\\seating.xlsx")
    sheets = wb.sheetnames
    for snname in file1.sheet_names:
        if snname == nm:
            c = file1.parse(j)
            s1 = wb[sheets[j]]
            ZZ = c.shape[1]
            break
        j += 1

    i = 0
    Label(window, text="", bg="lightblue2").pack(pady=20)
    canvas = Canvas(window, bg="lightblue2", width=500, height=800, highlightbackground="lightblue2", relief="flat")
    scrollbar = Scrollbar(canvas, orient=VERTICAL, command=canvas.yview)
    frame = Frame(canvas, bg="lightblue2", relief="flat")
    frame.grid(row=0, column=0)
    ws = window.winfo_screenwidth()
    hs = window.winfo_screenheight()
    w = 1000
    h = 1000
    x = (ws / 2) - (w / 2)
    y = (hs / 2) - (h / 2)
    Label(frame, text="NAME", font=("britannic bold", 20) ,height=1, width=15, bg="snow").grid(row=0, column=1)
    Label(frame, text="MARKS", font=("britannic bold", 20) ,height=1, width=15, bg="snow").grid(row=0, column=3)
    for index, row in c.iterrows():
        # label = Label(frame, text=row['Name'], font=16, bg="lightblue2").place(x=0, y=30 + i * 30)
        Label(window, text=str(s1.cell(index + 2, 1).value).upper() + " : ", font=20, bg="lightblue2").place(x=10,
                                                                                                             y=120 + i * 30)
        Label(window, text=str(s1.cell(index + 2, 5).value), font=20, bg="lightblue2").place(x=250, y=120 + i * 30)

        i = i + 1

    Button(window, text="BACK", command=viewA, bg="deepskyblue3").place(x=0, y=0)
    canvas.create_window(0, 0, anchor='nw', window=frame)
    canvas.update_idletasks()
    canvas.configure(scrollregion=canvas.bbox('all'), yscrollcommand=scrollbar.set)
    canvas.pack(fill='both', expand=True, side='left')
    scrollbar.pack(fill='y', side='right')


def viewA():
    clear()
    Button(window, text="BACK", command=welcomeA, font=4,height=1, bg="deepskyblue3",relief="groove").place(x=0,y=0)
    Label(window, text="", bg="lightblue2").pack(pady=20)
    course = []
    for sname in file1.sheet_names:
        course.append(sname)
    comboExample = ttk.Combobox(window,values=course,state='readonly',width=20,font=20,textvariable=v).pack(padx=10, pady=10)
    Button(window, text="View Marks",  font=16, bg="deepskyblue3", command=lambda:viewmarksbyadmin(v.get()),relief="solid", ).pack(padx=10, pady=10)
    Button(window, text="View Attendance", font=(16), bg="deepskyblue3", relief="solid",  command=lambda:attbyadmin(v.get())).pack(padx=10, pady=10)
    Button(window, text="View All Students", font=(16), bg="deepskyblue3", relief="solid", command=viewAllStudents).pack(padx=10, pady=10)

def newentry(nm):
    j = 0
    for snname in file1.sheet_names:
        if snname == nm:
            c = file1.parse(j)
            break
        j += 1
    i = 0
    count_row=c.shape[0]
    wb = load_workbook("C:\\Users\\USER\\Desktop\\seating.xlsx")
    sheets = wb.sheetnames
    s1 = wb[sheets[j]]
    s1.cell(count_row+2,1).value = name.get()
    s1.cell(count_row+2,3).value = regno.get()
    s1.cell(count_row+2,2).value = emailid.get()
    s1.cell(count_row+2,4).value = dept.get()
    wb.save("C:\\Users\\USER\\Desktop\\seating.xlsx")
    clear()
    Button(window, text="BACK", command=removestudent, font=4, height=1, bg="deepskyblue3", relief="groove").place(x=0,y=0)
    Label(window,text=name.get()+"'s Record Added Successfully",font=("microsoft tai le",20),bg="lightblue2",foreground="darkblue").pack(pady=40)

def addstudent():
    clear()
    Button(window, text="BACK", command=welcomeA, font=4,height=1, bg="deepskyblue3",relief="groove").place(x=0,y=0)

    Label(window,text="Enter Name:",bg="deepskyblue3",font=("book antiqua",12),relief="ridge",highlightthickness=5,highlightbackground='#222').place(x=50,y=50)
    Entry(window,textvariable=name,font=20).place(x=350,y=50)

    Label(window,text="Enter Registration number:",bg="deepskyblue3",font=("book antiqua",12),relief="ridge",highlightthickness=5,highlightbackground='#222').place(x=50,y=100)
    Entry(window,textvariable=regno,font=20).place(x=350,y=100)

    Label(window,text="Enter Email:",bg="deepskyblue3",font=("book antiqua",12),relief="ridge",highlightthickness=5,highlightbackground='#222').place(x=50,y=150)
    Entry(window,textvariable=emailid,font=20).place(x=350,y=150)

    Label(window,text="Enter Department:",bg="deepskyblue3",font=("book antiqua",12),relief="ridge",highlightthickness=5,highlightbackground='#222').place(x=50,y=200)
    Entry(window,textvariable=dept,font=20).place(x=350,y=200)
    Label(window, text="Choose Class:", bg="deepskyblue3", font=("book antiqua", 12), relief="ridge",highlightthickness=5, highlightbackground='#222').place(x=50, y=250)
    Button(window, text="ADD RECORD",font=5,height=1, bg="deepskyblue3",relief="groove",command=lambda:newentry(v.get())).place(x=50,y=350)
    course = []
    for sname in file1.sheet_names:
        course.append(sname)
    comboExample = ttk.Combobox(window, values=course, state='readonly',width=20,font=20, textvariable=v).place(x=350,y=250)

def Suspended(lbox,x2):
    x1=lbox.curselection()[0]
    wb=load_workbook("C:\\Users\\ASUS\\Desktop\\seating.xlsx")
    sheets=wb.sheetnames
    s1=wb[sheets[x2]]
    stuname=s1.cell(x1+2,1).value
    s1.delete_rows(x1+2,1)
    wb.save("C:\\Users\\ASUS\\Desktop\\seating.xlsx")
    clear()
    Button(window, text="BACK", command=removestudent, font=4, height=1, bg="deepskyblue3", relief="groove").place(x=0,y=0)
    Label(window, text=stuname+ "'s Record Deleted",bg="lightblue2",font=("microsoft tai le",20),foreground="darkblue").pack(pady=50)

def deleterecord(nm):
    clear()
    j=0
    Button(window, text="BACK", command=removestudent, font=4,height=1, bg="deepskyblue3",relief="groove").place(x=0, y=0)
    for snname in file1.sheet_names:
        if snname == nm:
            c = file1.parse(j)
            break
        j += 1
    Button(window, text="Delete the selected student",font=(16), bg="deepskyblue3", relief="solid", command=lambda: Suspended(lbox,j)).pack(pady=20)
    frame = Frame(window)
    frame.pack()
    lbox=Listbox(frame,width=50, height=30,bg="lightblue1",font=("century schoolbook", 12),relief="solid")
    yscroll = Scrollbar(frame,command=lbox.yview, orient=VERTICAL)
    yscroll.pack(side="right", fill="y")
    lbox.configure(yscrollcommand=yscroll.set)
    lbox.pack(side="left", fill="y")
    i=1
    for index, row in c.iterrows():
        lbox.insert(END,row['Name'])
        i = i + 1

def removestudent():
    clear()
    Button(window, text="BACK", command=welcomeA,font=4,height=1, bg="deepskyblue3",relief="groove").place(x=0, y=0)
    Label(window, text="", bg="lightblue2").pack(pady=20)
    Label(window, text=" Select a class from which to delete a student",font=("franklin gothic demi cond",12), bg="lightblue2").pack(pady=42)
    course = []
    for sname in file1.sheet_names:
        course.append(sname)
    comboExample = ttk.Combobox(window, values=course,state='readonly', width=20,font=20, textvariable=v).pack(pady=20)
    Button(window, text="CHOOSE CLASS", font=(16), bg="deepskyblue3", relief="solid",command=lambda: deleterecord(v.get())).pack()

def welcomeA():
    if username.get()=='AYA' and password.get()=='AYA':
        clear()
        Button(window, text="BACK", command=LoginA,font=6, bg="deepskyblue3",relief="groove").place(x=0,y=0)
        Label(window, text="", bg="lightblue2").pack(pady=10)
        Label(window,text='Welcome ADMIN', bg="deepskyblue3",font=("britannic bold",25),relief="solid",highlightthickness=16,highlightbackground='#222').pack(pady=10)
        Label(window, text="", bg="lightblue2").pack(pady=5)
        Button(window, text="VIEW", command=viewA,font=16, bg="deepskyblue3").pack(padx=10, pady=10)
        Button(window, text="ADD STUDENT", command=addstudent, font=16,bg="deepskyblue3").pack(padx=10, pady=10)
        Button(window, text="REMOVE  STUDENT", command=removestudent, font=16,bg="deepskyblue3").pack(padx=10, pady=10)

    else:
        Label(window, text="*INCORRECT CREDENTIALS", bg="snow",foreground="red").pack(pady=20)

def LoginA():
    clear()
    username.set("")
    password.set("")
    Button(window, text="BACK", font=5,height=1, bg="deepskyblue3",relief="groove", command=main).place(x=0,y=0)
    Label(window, text="", bg="lightblue2").pack(pady=20)
    Label(window,text="Enter username",bg="deepskyblue3",font=("book antiqua",16),relief="ridge",highlightthickness=5,highlightbackground='#222').pack(padx=10,pady=15)
    Entry(window,textvariable=username,bd=4,font=20).pack()
    Label(window,text="Enter password",bg="deepskyblue3",font=("book antiqua",16),relief="ridge",highlightthickness=5,highlightbackground='#222').pack(padx=10,pady=15)
    Entry(window,show="*",textvariable=password,bd=4,font=20).pack()
    Button(window, text ="LOGIN",font=16,bg="deepskyblue3",padx=5,pady=5,bd=5, fg="black",command=welcomeA).pack(padx=10,pady=15)

def LoginT():
    clear()
    username.set("")
    password.set("")
    Button(window, text="BACK", font=16, command=main,relief="groove", bg="deepskyblue3").place(x=0,y=0)
    Label(window, text="", bg="lightblue2").pack(pady=20)
    Label(window,text="Enter username",bd=4,bg="deepskyblue3",font=("book antiqua",16),relief="ridge",highlightthickness=5,highlightbackground='#222').pack(padx=10,pady=15)
    Entry(window,textvariable=username,bd=4,font=20).pack()
    Label(window,text="Enter password",bg="deepskyblue3",font=("book antiqua",16),relief="ridge",highlightthickness=5,highlightbackground='#222').pack(padx=10,pady=15)
    Entry(window,show="*",textvariable=password,bd=4,font=20).pack()
    Button(window, text ="LOGIN",font=16,bg="deepskyblue3",padx=5,pady=5,bd=5, fg="black",command=welcomeT).pack(padx=10,pady=15)

def LoginS():
    clear()
    username.set("")
    password.set("")
    Button(window, text="BACK", font=16, bg="deepskyblue3",relief="groove", command=main).place(x=0,y=0)
    Label(window, text="", bg="lightblue2").pack(pady=20)
    Label(window,text="Enter username",bg="deepskyblue3",font=("book antiqua",16),relief="ridge",highlightthickness=5,highlightbackground='#222').pack(padx=10,pady=15)
    Entry(window,textvariable=username,bd=4,font=20).pack()
    Label(window,text="Enter password",bg="deepskyblue3",font=("book antiqua",16),relief="ridge",highlightthickness=5,highlightbackground='#222').pack(padx=10,pady=15)
    Entry(window,show="*",font=20,bd=4,textvariable=password).pack()
    Button(window, text ="LOGIN",font=16,bg="deepskyblue3",padx=5,pady=5,bd=5, fg="black",command=welcomeS).pack(padx=10,pady=15)

def main():
    clear()
    Label(window,text="TEACHERS TOOL KIT",bg="deepskyblue3",font=("britannic bold",25),bd=4,relief="solid",highlightthickness=16,highlightbackground='#222').pack(padx=20,pady=20)
    Label(window,text="",bg="lightblue2").pack(pady=10)
    x=Button(window,text ="ADMIN LOGIN", padx=5,pady=5,bd=5, fg="black",font=16,command = LoginA,bg="deepskyblue3").pack(padx=10,pady=10)
    x = Button(window, text="TEACHER LOGIN",padx=5,pady=5,bd=5, fg="black",font=16, command=LoginT, bg="deepskyblue3").pack(padx=10,pady=10)
    x = Button(window, text="STUDENT LOGIN",padx=5,pady=5,bd=5, fg="black",font=16, command=LoginS, bg="deepskyblue3").pack(padx=10,pady=10)

main()
window.mainloop()
